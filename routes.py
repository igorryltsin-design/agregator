from flask import Blueprint, jsonify, request, Response, g, abort
from io import StringIO, BytesIO
import csv
from pathlib import Path
from datetime import datetime, date
import json

from models import (File, Tag, db, upsert_tag, file_to_dict, ChangeLog,
                    Collection, CollectionMember, User, UserActionLog,
                    AiSearchSnippetCache, AiSearchKeywordFeedback)
from flask import current_app
from sqlalchemy import func
import re

# --- Вспомогательные функции Natasha/pymorphy2 для морфологии и синонимов ---
try:
    from razdel import tokenize as _ru_tokenize
except Exception:
    _ru_tokenize = None
try:
    import pymorphy2
    _morph = pymorphy2.MorphAnalyzer()
except Exception:
    _morph = None

_RU_SYNONYMS = {
    'машина': ['автомобиль', 'авто', 'тачка', 'car'],
    'изображение': ['картинка', 'фото', 'фотография', 'image', 'picture', 'снимок', 'скриншот', 'screenshot'],
    'вуз': ['университет', 'институт', 'универ'],
    'статья': ['публикация', 'paper', 'publication', 'пейпер'],
    'диссертация': ['дисс', 'thesis', 'dissertation'],
    'журнал': ['journal', 'magazine', 'сборник'],
    'номер': ['выпуск', 'issue'],
    'страница': ['стр', 'страницы', 'pages', 'page'],
    'год': ['лет', 'г'],
}

_XLSX_ILLEGAL_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def _xlsx_safe(value):
    if value is None:
        return None
    if isinstance(value, (int, float, bool, datetime, date)):
        return value
    text = str(value)
    if not text:
        return ""
    sanitized = _XLSX_ILLEGAL_RE.sub(' ', text)
    def _valid_char(ch: str) -> bool:
        cp = ord(ch)
        if cp in (0x09, 0x0A, 0x0D):
            return True
        if 0x20 <= cp <= 0xD7FF:
            return True
        if 0xE000 <= cp <= 0xFFFD:
            return cp not in (0xFFFE, 0xFFFF)
        if 0x10000 <= cp <= 0x10FFFF:
            return True
        return False
    sanitized = ''.join(ch if _valid_char(ch) else ' ' for ch in sanitized)
    max_len = 32767
    if len(sanitized) > max_len:
        sanitized = sanitized[:max_len]
    return sanitized

def _ru_tokens(text: str) -> list[str]:
    s = (text or '').lower()
    if _ru_tokenize:
        return [t.text for t in _ru_tokenize(s)]
    # резерв: простое разбиение
    return re.sub(r"[^\w\d]+", " ", s).strip().split()

def _lemma(word: str) -> str:
    if _morph:
        try:
            p = _morph.parse(word)
            if p:
                return p[0].normal_form
        except Exception:
            pass
    return word.lower()

def _lemmas(text: str) -> list[str]:
    return [_lemma(w) for w in _ru_tokens(text)]

def _expand_synonyms(lemmas: list[str]) -> set[str]:
    out = set(lemmas)
    for l in list(lemmas):
        for s in _RU_SYNONYMS.get(l, []):
            out.add(_lemma(s))
    return out


def _normalize_tag_value(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    return re.sub(r"\s+", " ", raw).lower()


routes = Blueprint('routes', __name__)

def _current_user():
    return getattr(g, 'current_user', None)


def _require_admin():
    user = _current_user()
    if not user or getattr(user, 'role', '') != 'admin':
        abort(403)
    return user


def _allowed_collection_ids():
    return getattr(g, 'allowed_collection_ids', set())


def _apply_collection_filter(query, collection_model=None):
    allowed = _allowed_collection_ids()
    if allowed is None:
        return query
    model = collection_model or Collection
    if not allowed:
        return query.filter(model.id == -1)
    return query.filter(model.id.in_(allowed))


def _log_user_action(action: str, entity: str | None = None, entity_id: int | None = None, detail: str | None = None):
    user = _current_user()
    try:
        rec = UserActionLog(user_id=user.id if user else None,
                            action=action,
                            entity=entity,
                            entity_id=entity_id,
                            detail=detail)
        db.session.add(rec)
        db.session.commit()
    except Exception:
        db.session.rollback()


def _can_write_collection(collection_id: int | None) -> bool:
    user = _current_user()
    if not user:
        return False
    if getattr(user, 'role', '') == 'admin':
        return True
    if collection_id is None:
        return False
    try:
        col = Collection.query.get(collection_id)
        if col and col.owner_id == user.id:
            return True
    except Exception:
        pass
    try:
        member = CollectionMember.query.filter_by(collection_id=collection_id, user_id=user.id).first()
        if member and member.role in ('owner', 'editor'):
            return True
    except Exception:
        pass
    return False


def _require_collection_write(collection_id: int | None):
    if not _can_write_collection(collection_id):
        abort(403)


def _resolve_trash_directory(base_dir: Path | None = None) -> Path | None:
    base_dir = base_dir or Path(current_app.config.get('UPLOAD_FOLDER') or '.').expanduser()
    try:
        base_dir = base_dir.resolve()
    except Exception:
        pass
    deleted_cfg = current_app.config.get('DELETED_FOLDER')
    try:
        trash_dir = Path(deleted_cfg).expanduser() if deleted_cfg else Path('_deleted')
    except Exception:
        trash_dir = Path('_deleted')
    if not trash_dir.is_absolute():
        try:
            trash_dir = (base_dir / trash_dir).resolve()
        except Exception:
            trash_dir = base_dir / trash_dir
    else:
        try:
            trash_dir = trash_dir.resolve()
        except Exception:
            pass
    return trash_dir


def _collection_to_dict(col: Collection, include_members: bool = False) -> dict:
    info = {
        'id': col.id,
        'name': col.name,
        'slug': col.slug,
        'searchable': bool(col.searchable),
        'graphable': bool(col.graphable),
        'owner_id': col.owner_id,
        'owner_username': col.owner.username if getattr(col, 'owner', None) else None,
        'is_private': bool(getattr(col, 'is_private', False)),
        'created_at': col.created_at.isoformat() if getattr(col, 'created_at', None) else None,
        'count': len(col.files or []),
    }
    if include_members:
        members = []
        try:
            for m in col.members or []:
                members.append({
                    'user_id': m.user_id,
                    'role': m.role,
                    'username': m.user.username if m.user else None,
                })
        except Exception:
            members = []
        info['members'] = members
    return info


def _require_collection_owner_or_admin(col: Collection):
    user = _current_user()
    if not user:
        abort(403)
    if getattr(user, 'role', '') == 'admin':
        return user
    if col.owner_id == user.id:
        return user
    member = CollectionMember.query.filter_by(collection_id=col.id, user_id=user.id).first()
    if member and member.role == 'owner':
        return user
    abort(403)

@routes.route("/api/files", methods=["GET"])
def api_files():
    user = _current_user()
    if not user:
        abort(401)
    try:
        files = File.query.join(Collection, File.collection_id == Collection.id) \
            .filter(Collection.searchable == True)
        files = _apply_collection_filter(files, Collection) \
            .order_by(File.mtime.desc().nullslast()).limit(200).all()
    except Exception:
        q = File.query
        allowed = _allowed_collection_ids()
        if allowed is not None:
            if not allowed:
                q = q.filter(File.collection_id == -1)
            else:
                q = q.filter(File.collection_id.in_(allowed))
        files = q.order_by(File.mtime.desc().nullslast()).limit(200).all()
    return jsonify([file_to_dict(f) for f in files])

@routes.route("/api/files/<int:file_id>", methods=["GET"])
def api_file_detail(file_id):
    f = File.query.get_or_404(file_id)
    allowed = _allowed_collection_ids()
    if allowed is not None and f.collection_id not in allowed:
        abort(403)
    return jsonify(file_to_dict(f))


@routes.route('/api/graph')
def api_graph():
    if not _current_user():
        abort(401)
    """Граф по выбранным тегам.
    Query:
      - keys: список ключей тегов через запятую (напр. author,advisor,organization). Если пусто — используется author.
      - limit: макс. файлов (по умолчанию 500) для ограничения размера графа.
    Узлы: file:<id> (type=work) и tag:<key>:<value> (type=tag:key)
    Рёбра: file -> tag (label=key)
    """
    keys_param = (request.args.get('keys') or '').strip()
    tag_keys = [k.strip() for k in keys_param.split(',') if k.strip()] or ['author']
    allowed_cfg = current_app.config.get('GRAPH_FACET_TAG_KEYS')
    if isinstance(allowed_cfg, (list, tuple)):
        allowed_keys = [str(v or '').strip() for v in allowed_cfg if str(v or '').strip()]
    elif allowed_cfg is None:
        allowed_keys = None
    else:
        allowed_keys = []
    if allowed_keys is not None:
        tag_keys = [k for k in tag_keys if k in allowed_keys]
        if not tag_keys:
            tag_keys = allowed_keys[:3] if allowed_keys else []
    if not tag_keys:
        tag_keys = ['author']
    try:
        limit = max(min(int(request.args.get('limit', '500')), 2000), 1)
    except Exception:
        limit = 500
    q = File.query.join(Collection, File.collection_id == Collection.id)
    q = _apply_collection_filter(q, Collection)
    try:
        q = q.filter(Collection.graphable == True)
    except Exception:
        pass
    # фильтры по годам (строковое поле сохранено для совместимости)
    year_from = (request.args.get('year_from') or '').strip()
    year_to = (request.args.get('year_to') or '').strip()
    if year_from:
        q = q.filter(File.year >= year_from)
    if year_to:
        q = q.filter(File.year <= year_to)
    files = q.order_by(File.mtime.desc().nullslast()).limit(limit).all()

    snapshot_fn = current_app.config.get('authority_snapshot_fn')
    authority_snapshot: dict = {}
    if callable(snapshot_fn):
        try:
            authority_snapshot = snapshot_fn(_allowed_collection_ids())
        except Exception as exc:
            current_app.logger.debug('authority snapshot failed: %s', exc)
            authority_snapshot = {}
    doc_scores = authority_snapshot.get('doc_scores', {}) if authority_snapshot else {}
    topic_index = authority_snapshot.get('topic_index', {}) if authority_snapshot else {}
    author_entries = authority_snapshot.get('author_entries', []) if authority_snapshot else []
    author_index = {
        _normalize_tag_value(entry.get('name')): float(entry.get('score') or 0.0)
        for entry in author_entries
        if isinstance(entry, dict) and entry.get('name')
    }

    nodes: list[dict] = []
    edges: list[dict] = []
    file_ids = set()
    tag_node_ids: dict[tuple[str, str], str] = {}
    next_tag_id = 100000
    max_authority = 0.0
    for f in files:
        fid = f.id
        if fid not in file_ids:
            authority_score = float(doc_scores.get(fid, 0.0))
            nodes.append({
                "id": f"file-{fid}",
                "label": f.title or f.filename or str(fid),
                "type": "work",
                "authority": authority_score,
            })
            file_ids.add(fid)
            if authority_score > max_authority:
                max_authority = authority_score
        # Собираем подходящие теги
        for t in (f.tags or []):
            if t.key in tag_keys and (t.value or '').strip():
                key = t.key
                val = t.value.strip()
                norm_val = _normalize_tag_value(val)
                tag_key = (key, norm_val or val)
                nid = tag_node_ids.get(tag_key)
                if not nid:
                    nid = f"tag-{key}-{next_tag_id}"
                    tag_node_ids[tag_key] = nid
                    authority_score = float(topic_index.get(f"{key}|||{norm_val}", 0.0))
                    nodes.append({
                        "id": nid,
                        "label": val,
                        "type": f"tag:{key}",
                        "authority": authority_score,
                    })
                    next_tag_id += 1
                    if authority_score > max_authority:
                        max_authority = authority_score
                edges.append({"from": f"file-{fid}", "to": nid, "label": key})
        # Резервный вариант по полям File (например author)
        if 'author' in tag_keys and (f.author or '').strip():
            val = f.author.strip()
            norm_val = _normalize_tag_value(val)
            tag_key = ('author', norm_val or val)
            nid = tag_node_ids.get(tag_key)
            if not nid:
                nid = f"tag-author-{next_tag_id}"
                tag_node_ids[tag_key] = nid
                authority_score = float(topic_index.get(f"author|||{norm_val}", 0.0))
                if not authority_score and norm_val:
                    authority_score = float(author_index.get(norm_val, 0.0))
                nodes.append({
                    "id": nid,
                    "label": val,
                    "type": "tag:author",
                    "authority": authority_score,
                })
                next_tag_id += 1
                if authority_score > max_authority:
                    max_authority = authority_score
            edges.append({"from": f"file-{fid}", "to": nid, "label": 'author'})
    # Необязательный умный поиск на сервере с русскими леммами и синонимами
    q_str = (request.args.get('q') or '').strip()
    smart = str(request.args.get('smart', '')).lower() in ('1','true','yes','on')
    if q_str and smart:
        q_lemmas = list(_expand_synonyms(_lemmas(q_str)))
        def match(label: str) -> bool:
            lab = set(_lemmas(label or ''))
            return any(l in lab for l in q_lemmas)
        keep_ids = set()
        for n in nodes:
            if match(n.get('label') or ''):
                keep_ids.add(n['id'])
        # Оставляем рёбра только когда обе вершины сохранены; расширяем множество, чтобы включить соседей и не изолировать совпадения
        neigh_keep = set(keep_ids)
        for e in edges:
            if e['from'] in keep_ids or e['to'] in keep_ids:
                neigh_keep.add(e['from']); neigh_keep.add(e['to'])
        nodes = [n for n in nodes if n['id'] in neigh_keep]
        kept = set(n['id'] for n in nodes)
        edges = [e for e in edges if e['from'] in kept and e['to'] in kept]
        max_authority = max((float(n.get('authority', 0.0)) for n in nodes), default=0.0)
    else:
        max_authority = max((float(n.get('authority', 0.0)) for n in nodes), default=max_authority)
    return jsonify({"nodes": nodes, "edges": edges, "keys": tag_keys, "max_authority": max_authority})


@routes.route('/api/graph/build', methods=['POST'])
def api_graph_build():
    """Создать недостающие теги для автора и научного руководителя.
    Это помогает заполнить связи, используемые в графе.
    """
    _require_admin()
    files = File.query.all()
    created = 0
    for f in files:
        # автор
        if f.author:
            if not any(t.key == 'author' and t.value == f.author for t in f.tags):
                upsert_tag(f, 'author', f.author)
                created += 1
        # научный руководитель -> тег, похожий на организацию
        if f.advisor:
            if not any(t.key == 'advisor' and t.value == f.advisor for t in f.tags):
                upsert_tag(f, 'advisor', f.advisor)
                created += 1
    db.session.commit()
    return jsonify({"created_tags": created})

@routes.route("/api/files", methods=["POST"])
def api_file_create():
    data = request.json or {}
    user = _current_user()
    if not user:
        abort(401)
    col_id = data.get('collection_id')
    if not col_id:
        try:
            personal = Collection.query.filter_by(owner_id=user.id, is_private=True).first()
            if personal:
                col_id = personal.id
        except Exception:
            col_id = None
        if not col_id:
            try:
                base_col = Collection.query.filter_by(slug='base').first()
                col_id = base_col.id if base_col else None
            except Exception:
                col_id = None
    _require_collection_write(col_id)
    file_path = Path(current_app.config.get('UPLOAD_FOLDER') or '.') / (data.get('filename') or '')
    f = File(
        title=data.get("title"),
        author=data.get("author"),
        year=data.get("year"),
        material_type=data.get("material_type"),
        filename=data.get("filename"),
        keywords=data.get("keywords"),
        collection_id=col_id,
        rel_path=data.get("filename"),
        path=str(file_path),
    )
    db.session.add(f)
    db.session.flush()
    for tag in data.get("tags", []):
        upsert_tag(f, tag.get("key"), tag.get("value"))
    db.session.flush()
    try:
        from app import _sync_file_to_fts
        _sync_file_to_fts(f)
    except Exception:
        pass
    db.session.commit()
    _log_user_action('file_create', 'file', f.id, detail=str(data))
    invalidate = current_app.config.get('invalidate_facets')
    if callable(invalidate):
        invalidate('file create')
    return jsonify(file_to_dict(f)), 201

@routes.route("/api/files/<int:file_id>", methods=["PUT"])
def api_file_update(file_id):
    f = File.query.get_or_404(file_id)
    _require_collection_write(f.collection_id)
    data = request.json or {}
    old_type = (f.material_type or '').strip().lower()
    actor = _current_user()
    for field in ["title", "author", "year", "material_type", "filename", "keywords"]:
        if field in data:
            setattr(f, field, data[field])
    Tag.query.filter_by(file_id=f.id).delete()
    for tag in data.get("tags", []):
        upsert_tag(f, tag.get("key"), tag.get("value"))
    # Автоперенос при изменении типа
    try:
        new_type = (f.material_type or '').strip().lower()
        move_enabled = bool(current_app.config.get('MOVE_ON_RENAME', True))
        if move_enabled and new_type and new_type != old_type:
            base_dir = Path(current_app.config.get('UPLOAD_FOLDER') or '.')
            type_dirs = current_app.config.get('TYPE_DIRS') or {}
            target_sub = type_dirs.get(new_type) or type_dirs.get('other') or 'other'
            target_dir = base_dir / target_sub
            target_dir.mkdir(parents=True, exist_ok=True)
            p_old = Path(f.path)
            ext = f.ext or p_old.suffix
            base = f.filename or p_old.stem
            p_new = target_dir / (base + (ext or ''))
            i = 1
            while p_new.exists() and p_new.resolve() != p_old.resolve():
                p_new = target_dir / (f"{base}_{i}" + (ext or ''))
                i += 1
            # переносим файл
            p_old.rename(p_new)
            # очищаем старую миниатюру для PDF
            try:
                if (ext or '').lower() == '.pdf':
                    thumb = Path(current_app.static_folder) / 'thumbnails' / (p_old.stem + '.png')
                    if thumb.exists(): thumb.unlink()
            except Exception:
                pass
            # обновляем поля в БД
            f.path = str(p_new)
            try:
                f.rel_path = str(p_new.relative_to(base_dir))
            except Exception:
                f.rel_path = p_new.name
            f.filename = p_new.stem
            try:
                f.mtime = p_new.stat().st_mtime
            except Exception:
                pass
            try:
                db.session.add(ChangeLog(file_id=f.id, action='move', field='material_type', old_value=old_type, new_value=new_type, info=f"{p_old} -> {p_new}"))
            except Exception:
                pass
    except Exception as e:
        current_app.logger.warning(f"Auto-move on type change failed: {e}")
    db.session.flush()
    try:
        from app import _sync_file_to_fts
        _sync_file_to_fts(f)
    except Exception:
        pass
    db.session.commit()
    try:
        tag_snapshot = []
        for tag in data.get("tags", []):
            k = (tag or {}).get('key')
            v = (tag or {}).get('value')
            if k and v:
                tag_snapshot.append(f"{k}={v}")
        current_app.logger.info(
            "[tags] user=%s file=%s updated fields=%s tags=%s",
            getattr(actor, 'username', None) or 'system',
            f.id,
            {k: data.get(k) for k in ("title", "author", "year", "material_type", "keywords") if k in data},
            '; '.join(tag_snapshot)
        )
    except Exception:
        pass
    _log_user_action('file_update', 'file', f.id, detail=str(data))
    invalidate = current_app.config.get('invalidate_facets')
    if callable(invalidate):
        invalidate('file update')
    return jsonify(file_to_dict(f))

@routes.route('/api/files/move-by-type', methods=['POST'])
def api_move_by_type():
    """Перенести группу файлов в подпапки по текущему типу.
    JSON: {"ids":[...]} или {"all":true}
    """
    _require_admin()
    data = request.json or {}
    ids = data.get('ids') or []
    move_all = bool(data.get('all'))
    base_dir = Path(current_app.config.get('UPLOAD_FOLDER') or '.')
    type_dirs = current_app.config.get('TYPE_DIRS') or {}
    moved = 0
    skipped = 0
    errors = []
    q = File.query
    if not move_all:
        if not ids:
            return jsonify({"ok": False, "error": "ids or all=true required"}), 400
        q = q.filter(File.id.in_(ids))
    files = q.all()
    for f in files:
        try:
            mt = (f.material_type or '').strip().lower()
            sub = type_dirs.get(mt) or type_dirs.get('other') or 'other'
            target_dir = base_dir / sub
            p_old = Path(f.path)
            if not p_old.exists():
                skipped += 1
                continue
            # пропускаем, если файл уже в целевой папке
            try:
                if target_dir.resolve() == p_old.parent.resolve():
                    skipped += 1
                    continue
            except Exception:
                pass
            target_dir.mkdir(parents=True, exist_ok=True)
            ext = f.ext or p_old.suffix
            base = f.filename or p_old.stem
            p_new = target_dir / (base + (ext or ''))
            i = 1
            while p_new.exists() and p_new.resolve() != p_old.resolve():
                p_new = target_dir / (f"{base}_{i}" + (ext or ''))
                i += 1
            p_old.rename(p_new)
            # очищаем старую миниатюру для PDF
            try:
                if (ext or '').lower() == '.pdf':
                    thumb = Path(current_app.static_folder) / 'thumbnails' / (p_old.stem + '.png')
                    if thumb.exists(): thumb.unlink()
            except Exception:
                pass
            # обновляем запись в БД
            f.path = str(p_new)
            try:
                f.rel_path = str(p_new.relative_to(base_dir))
            except Exception:
                f.rel_path = p_new.name
            f.filename = p_new.stem
            try:
                f.mtime = p_new.stat().st_mtime
            except Exception:
                pass
            try:
                db.session.add(ChangeLog(file_id=f.id, action='move', field='material_type', old_value=mt, new_value=mt, info=f"{p_old} -> {p_new}"))
            except Exception:
                pass
            moved += 1
        except Exception as e:
            errors.append(str(e))
    db.session.commit()
    return jsonify({"ok": True, "moved": moved, "skipped": skipped, "errors": errors})

@routes.route("/api/files/<int:file_id>", methods=["DELETE"])
def api_file_delete(file_id):
    f = File.query.get_or_404(file_id)
    _require_collection_write(f.collection_id)
    actor = _current_user()
    remove_fs = str(request.args.get('rm', '')).lower() in ('1', 'true', 'yes', 'on')

    # Сохраняем необходимые данные до удаления записи
    file_path = Path(f.path).expanduser() if f.path else None
    rel_path_raw = (f.rel_path or '').strip()
    sha1 = (f.sha1 or '').strip()

    # Чистим связанные записи вручную, чтобы избежать ошибок каскадного удаления в SQLite
    try:
        Tag.query.filter_by(file_id=f.id).delete(synchronize_session=False)
        ChangeLog.query.filter_by(file_id=f.id).delete(synchronize_session=False)
        AiSearchSnippetCache.query.filter_by(file_id=f.id).delete(synchronize_session=False)
        AiSearchKeywordFeedback.query.filter_by(file_id=f.id).update({AiSearchKeywordFeedback.file_id: None}, synchronize_session=False)
        db.session.delete(f)
        db.session.commit()
        try:
            from app import _delete_file_from_fts
            _delete_file_from_fts(file_id)
        except Exception:
            pass
    except Exception as exc:
        db.session.rollback()
        current_app.logger.error(f"Failed to delete file record {file_id}: {exc}", exc_info=True)
        return jsonify({'ok': False, 'error': 'Не удалось удалить запись из базы'}), 500

    warnings: list[str] = []
    moved_to: str | None = None

    # Работа с файловой системой выполняется после успешного удаления из БД
    if file_path and file_path.exists():
        if remove_fs:
            try:
                file_path.unlink()
            except Exception as exc:
                current_app.logger.warning(f"Failed to delete file on disk: {exc}")
                warnings.append(f"fs:{exc}")
        else:
            try:
                base_dir = Path(current_app.config.get('UPLOAD_FOLDER') or '.').expanduser()
                try:
                    base_dir = base_dir.resolve()
                except Exception:
                    pass
                trash_dir = _resolve_trash_directory(base_dir) or (base_dir / '_deleted')
                try:
                    trash_dir.mkdir(parents=True, exist_ok=True)
                except Exception:
                    pass

                rel_candidate: Path | None = None
                if rel_path_raw:
                    sanitized = rel_path_raw.replace('\\', '/').lstrip('/').strip()
                    rel_parts = [p for p in sanitized.split('/') if p not in ('', '.', '..')]
                    if rel_parts:
                        rel_candidate = Path(*rel_parts)
                if rel_candidate is None:
                    try:
                        rel_candidate = file_path.relative_to(base_dir)
                    except Exception:
                        rel_candidate = Path(file_path.name)

                target_path = trash_dir / rel_candidate
                target_path.parent.mkdir(parents=True, exist_ok=True)
                if target_path.exists():
                    suffix = target_path.suffix
                    stem = target_path.stem
                    ts = datetime.utcnow().strftime('%Y%m%d%H%M%S')
                    target_path = target_path.with_name(f"{stem}_{ts}{suffix}")
                file_path.rename(target_path)
                moved_to = str(target_path)
            except Exception as exc:
                current_app.logger.warning(f"Failed to move file to trash: {exc}")
                warnings.append(f"move:{exc}")
    elif file_path:
        warnings.append('fs:missing')

    # Удаляем производные артефакты (thumbnail, текстовые фрагменты)
    rel_path_for_artifacts = rel_path_raw
    try:
        thumb_base = Path(rel_path_for_artifacts).stem if rel_path_for_artifacts else None
        if thumb_base:
            thumb = Path(current_app.static_folder) / 'thumbnails' / (thumb_base + '.png')
            if thumb.exists():
                thumb.unlink()
    except Exception as exc:
        current_app.logger.warning(f"Failed to delete thumbnail: {exc}")
        warnings.append(f"thumb:{exc}")

    try:
        cache_dir = Path(current_app.static_folder) / 'cache' / 'text_excerpts'
        safe_rel = rel_path_for_artifacts.replace('/', '_').replace('\\', '_') if rel_path_for_artifacts else ''
        key_base = sha1 or safe_rel
        if key_base:
            cache_file = cache_dir / f"{key_base}.txt"
            if cache_file.exists():
                cache_file.unlink()
    except Exception as exc:
        current_app.logger.warning(f"Failed to delete cached excerpt: {exc}")
        warnings.append(f"cache:{exc}")

    detail_parts = [f"rm={int(remove_fs)}", f"user={getattr(actor, 'id', None)}"]
    if moved_to:
        detail_parts.append(f"moved_to={moved_to}")
    if warnings:
        detail_parts.append(f"warnings={';'.join(warnings)}")
    _log_user_action('file_delete', 'file', file_id, detail='; '.join(detail_parts))
    invalidate = current_app.config.get('invalidate_facets')
    if callable(invalidate):
        invalidate('file delete')
    return "", 204

@routes.route('/api/admin/collections', methods=['GET'])
def api_admin_collections():
    _require_admin()
    cols = Collection.query.order_by(Collection.name.asc()).all()
    return jsonify({'ok': True, 'collections': [_collection_to_dict(c, include_members=True) for c in cols]})


@routes.route('/api/collections/<int:collection_id>/members', methods=['GET', 'POST'])
def api_collection_members(collection_id: int):
    col = Collection.query.get_or_404(collection_id)
    actor = _require_collection_owner_or_admin(col)
    if request.method == 'GET':
        return jsonify({'ok': True, 'members': _collection_to_dict(col, include_members=True)['members']})
    data = request.get_json(silent=True) or {}
    role = str(data.get('role') or 'viewer').strip().lower()
    if role not in ('viewer', 'editor', 'owner'):
        role = 'viewer'
    raw_identifier = data.get('user_id')
    username_hint = data.get('username')
    user_obj: User | None = None
    user_id: int | None = None

    def _lookup_by_username(value: str | None) -> User | None:
        if not value:
            return None
        login = str(value).strip()
        if not login:
            return None
        return User.query.filter(func.lower(User.username) == login.lower()).first()

    try:
        if raw_identifier is not None:
            if isinstance(raw_identifier, int):
                user_obj = User.query.get(raw_identifier)
            else:
                raw_str = str(raw_identifier).strip()
                if raw_str.isdigit():
                    user_obj = User.query.get(int(raw_str))
                if not user_obj:
                    user_obj = _lookup_by_username(raw_str)
        if not user_obj:
            user_obj = _lookup_by_username(username_hint)
    except Exception:
        user_obj = None

    if not user_obj:
        return jsonify({'ok': False, 'error': 'Пользователь не найден'}), 404

    user_id = user_obj.id
    if role == 'owner' and getattr(actor, 'role', '') != 'admin':
        role = 'editor'
    member = CollectionMember.query.filter_by(collection_id=collection_id, user_id=user_id).first()
    if member:
        member.role = role
    else:
        member = CollectionMember(collection_id=collection_id, user_id=user_id, role=role)
        db.session.add(member)
    db.session.commit()
    detail = json.dumps({'user_id': user_id, 'username': user_obj.username, 'role': role})
    _log_user_action('collection_member_add', 'collection', collection_id, detail=detail)
    col = Collection.query.get(collection_id)
    return jsonify({'ok': True, 'members': _collection_to_dict(col, include_members=True)['members']})


@routes.route('/api/collections/<int:collection_id>/members/<int:user_id>', methods=['PATCH', 'DELETE'])
def api_collection_member_detail(collection_id: int, user_id: int):
    col = Collection.query.get_or_404(collection_id)
    actor = _require_collection_owner_or_admin(col)
    member = CollectionMember.query.filter_by(collection_id=collection_id, user_id=user_id).first()
    if not member:
        abort(404)
    if request.method == 'DELETE':
        username = member.user.username if member.user else None
        if user_id == col.owner_id and getattr(actor, 'role', '') != 'admin':
            return jsonify({'ok': False, 'error': 'Нельзя удалить владельца коллекции'}), 400
        db.session.delete(member)
        db.session.commit()
        detail = json.dumps({'user_id': user_id, 'username': username})
        _log_user_action('collection_member_remove', 'collection', collection_id, detail=detail)
        return jsonify({'ok': True})
    data = request.get_json(silent=True) or {}
    role = str(data.get('role') or '').strip().lower()
    if role in ('viewer', 'editor', 'owner'):
        if role == 'owner' and getattr(actor, 'role', '') != 'admin':
            role = 'editor'
        if user_id == col.owner_id and role != 'owner' and getattr(actor, 'role', '') != 'admin':
            return jsonify({'ok': False, 'error': 'Нельзя менять роль владельца'}), 400
        member.role = role
        db.session.commit()
        username = member.user.username if member.user else None
        detail = json.dumps({'user_id': user_id, 'username': username, 'role': role})
        _log_user_action('collection_member_update', 'collection', collection_id, detail=detail)
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'Некорректная роль'}), 400

@routes.route("/export/csv")
def export_csv():
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Name', 'Tags'])
    files = File.query.all()
    for file in files:
        writer.writerow([file.id, file.filename, ', '.join(f"{t.key}={t.value}" for t in file.tags)])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment;filename=export.csv"})

@routes.route("/export/bibtex")
def export_bibtex():
    output = StringIO()
    files = File.query.all()
    for file in files:
        output.write(f"@misc{{{file.id},\n  title={{ {file.filename} }},\n  tags={{ {', '.join(f'{t.key}={t.value}' for t in file.tags)} }}\n}}\n")
    return Response(output.getvalue(), mimetype="text/x-bibtex", headers={"Content-Disposition": "attachment;filename=export.bib"})


@routes.route("/api/collections/<int:collection_id>/export/excel")
def export_collection_excel(collection_id: int):
    allowed = _allowed_collection_ids()
    if allowed is not None and collection_id not in allowed:
        abort(403)

    collection = Collection.query.get_or_404(collection_id)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        current_app.logger.error("Excel export unavailable: openpyxl is not installed")
        return jsonify({'ok': False, 'error': 'Экспорт в Excel недоступен: требуется openpyxl'}), 500

    files = File.query.filter(File.collection_id == collection_id).order_by(File.id.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Files"

    headers = [
        "File",
        "Collection Relative Path",
        "Collection",
        "ID",
        "Title",
        "Author",
        "Year",
        "Material Type",
        "Advisor",
        "Keywords",
        "Abstract",
        "Filename",
        "Extension",
        "Size (bytes)",
        "Modified (UTC)",
        "SHA1",
        "Tags",
    ]
    ws.append(headers)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    slug = (collection.slug or '').strip()
    prefix = f"collections/{slug}/" if slug else None
    row_count = 0
    max_lengths = [len(h) for h in headers]

    for f in files:
        rel_path = (f.rel_path or '').replace('\\', '/')
        if prefix and rel_path.startswith(prefix):
            col_rel_path = rel_path[len(prefix):]
        else:
            col_rel_path = rel_path

        display_name = (f.title or f.filename or Path(rel_path).name or f"file-{f.id}").strip()

        tag_map: dict[str, list[str]] = {}
        for t in getattr(f, 'tags', []) or []:
            key = (getattr(t, 'key', '') or '').strip()
            value = (getattr(t, 'value', '') or '').strip()
            if not key or not value:
                continue
            tag_map.setdefault(key, []).append(value)
        tag_repr = '; '.join(
            f"{k}={', '.join(sorted(set(vs)))}" for k, vs in sorted(tag_map.items())
        )

        modified = None
        if getattr(f, 'mtime', None):
            try:
                modified = datetime.utcfromtimestamp(float(f.mtime)).isoformat()
            except Exception:
                modified = str(f.mtime)

        row_raw = [
            display_name,
            col_rel_path,
            collection.name,
            f.id,
            f.title,
            f.author,
            f.year,
            f.material_type,
            f.advisor,
            f.keywords,
            f.abstract,
            f.filename,
            f.ext,
            f.size,
            modified,
            f.sha1,
            tag_repr,
        ]
        row = [_xlsx_safe(val) for val in row_raw]
        ws.append(row)
        row_idx = ws.max_row
        link_cell = ws.cell(row=row_idx, column=1)
        if col_rel_path:
            safe_link = _xlsx_safe(col_rel_path)
            link_cell.hyperlink = safe_link
            link_cell.style = "Hyperlink"

        for idx, val in enumerate(row, start=1):
            if val is None:
                continue
            length = len(str(val))
            if length > max_lengths[idx - 1]:
                max_lengths[idx - 1] = min(length, 80)

        row_count += 1

    for idx, width in enumerate(max_lengths, start=1):
        column_letter = get_column_letter(idx)
        ws.column_dimensions[column_letter].width = width + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    ts = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    filename = f"{slug or 'collection'}-{ts}.xlsx"

    try:
        detail = json.dumps({'files': row_count})
    except Exception:
        detail = None
    _log_user_action('collection_export_excel', 'collection', collection_id, detail=detail)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@routes.route("/api/aiword/bibtex")
def api_aiword_bibtex():
    """Сгенерировать BibTeX на основе каталога для AiWord.
    Пытаемся сопоставить material_type и теги стандартным полям BibTeX.
    """
    def tag_get(f: File, names: list[str]) -> str | None:
        want = {n.lower() for n in names}
        for t in (f.tags or []):
            if (t.key or '').lower() in want and (t.value or '').strip():
                return t.value.strip()
        return None

    def bib_for_file(f: File) -> str:
        mt = (f.material_type or '').strip().lower()
        key = f"f{f.id or 0}"
        # выбираем тип записи
        if mt == 'article':
            et = 'article'
        elif mt in ('textbook', 'monograph', 'book'):
            et = 'book'
        elif mt in ('inproceedings', 'proceedings'):
            et = 'inproceedings'
        elif mt in ('dissertation', 'dissertation_abstract'):
            # по умолчанию используем phdthesis
            deg = tag_get(f, ['degree', 'степень']) or ''
            et = 'phdthesis' if 'доктор' in deg.lower() else 'mastersthesis'
        else:
            et = 'misc'

        title = (f.title or f.filename or '').strip() or 'Untitled'
        author = (f.author or '').strip()
        year = (f.year or '').strip()
        # Распространённые теги
        doi = tag_get(f, ['doi'])
        pages = tag_get(f, ['pages', 'страницы'])
        journal = tag_get(f, ['journal', 'журнал'])
        vol_issue = tag_get(f, ['volume_issue'])
        publisher = tag_get(f, ['publisher', 'издательство'])
        school = tag_get(f, ['organization', 'организация'])
        # Формируем поля
        fields: list[str] = [f"title={{ {title} }}"]
        if author:
            fields.append(f"author={{ {author} }}")
        if year:
            fields.append(f"year={{ {year} }}")
        if et == 'article':
            if journal:
                fields.append(f"journal={{ {journal} }}")
            if vol_issue and '/' in vol_issue:
                v, n = vol_issue.split('/', 1)
                v = v.strip(); n = n.strip()
                if v:
                    fields.append(f"volume={{ {v} }}")
                if n:
                    fields.append(f"number={{ {n} }}")
            if pages:
                fields.append(f"pages={{ {pages} }}")
        elif et in ('book',):
            if publisher:
                fields.append(f"publisher={{ {publisher} }}")
            if pages:
                fields.append(f"pages={{ {pages} }}")
        elif et in ('inproceedings',):
            bt = tag_get(f, ['conference'])
            if bt:
                fields.append(f"booktitle={{ {bt} }}")
            if publisher:
                fields.append(f"publisher={{ {publisher} }}")
            if pages:
                fields.append(f"pages={{ {pages} }}")
        elif et in ('phdthesis','mastersthesis'):
            if school:
                fields.append(f"school={{ {school} }}")
            if pages:
                fields.append(f"pages={{ {pages} }}")
        if doi:
            fields.append(f"doi={{ {doi} }}")
        # Ссылка на страницу элемента внутри приложения
        try:
            fields.append(f"url={{ /file/{f.id} }}")
        except Exception:
            pass

        return f"@{et}{{{key},\n  " + ",\n  ".join(fields) + "\n}\n"

    # Отдаём предпочтение только коллекциям с поиском; при ошибке берём все
    try:
        q = File.query.join(Collection, File.collection_id == Collection.id).filter(Collection.searchable == True)
        q = _apply_collection_filter(q, Collection)
        files = q.all()
    except Exception:
        q = File.query
        allowed = _allowed_collection_ids()
        if allowed is not None:
            if not allowed:
                q = q.filter(File.collection_id == -1)
            else:
                q = q.filter(File.collection_id.in_(allowed))
        files = q.all()
    out = StringIO()
    for f in files:
        try:
            out.write(bib_for_file(f))
        except Exception:
            # резервный минимальный вариант записи misc
            out.write(f"@misc{{f{getattr(f,'id',0)}, title={{ {getattr(f,'title',None) or getattr(f,'filename','')} }} }}\n")
    return Response(out.getvalue(), mimetype="text/x-bibtex")

@routes.route('/api/collections')
def api_collections():
    allowed = _allowed_collection_ids()
    q = Collection.query.order_by(Collection.name.asc())
    if allowed is not None:
        if not allowed:
            q = q.filter(Collection.id == -1)
        else:
            q = q.filter(Collection.id.in_(allowed))
    cols = q.all()
    out = []
    for c in cols:
        out.append({
            'id': c.id,
            'name': c.name,
            'slug': c.slug,
            'searchable': bool(getattr(c, 'searchable', True)),
            'graphable': bool(getattr(c, 'graphable', True)),
            'count': len(c.files or []),
        })
    return jsonify(out)
